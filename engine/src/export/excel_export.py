from __future__ import annotations

from pathlib import Path
from typing import Optional, Union

from domain.models import ParsedOrder
from .utils import sanitize_filename


def export_to_excel(
    result: ParsedOrder,
    output_dir: Union[str, Path],
    filename: Optional[str] = None,
) -> str:
    """
    Export parsed order to Excel file.

    Args:
        result: ParsedOrder object
        output_dir: Directory to save file
        filename: Optional filename (auto-generated if not provided)

    Returns:
        Path to saved file (string)
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError as e:
        raise ImportError(
            "openpyxl is required for Excel export. Install with: poetry add openpyxl"
        ) from e

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    if filename is None:
        filename = f"{result.get_filename()}.xlsx"

    filename = sanitize_filename(filename)
    filepath = output_dir / filename

    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Failed to create worksheet in Excel workbook")

    # Generate tab name (Excel limit 31 chars)
    tab_name = result.get_filename()[:31]
    ws.title = tab_name

    # Styles
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    label_font = Font(bold=True)

    # =====================
    # Main Order Information
    # =====================
    ws["A1"] = "Main Order Information"
    ws["A1"].font = header_font
    ws["A1"].fill = header_fill
    ws.merge_cells("A1:D1")

    info = result.main_order_info
    fields = [
        ("Business Client", info.business_client),
        ("Client Name", info.client_name),
        ("Client Information", info.client_information),
        ("Order Subtotal", f"${info.order_subtotal:.2f}" if info.order_subtotal else ""),
        ("Requested Pickup Time", info.requested_pick_up_time),
        ("Requested Pickup Date", info.requested_pick_up_date),
        ("Number of Guests", info.number_of_guests),
        ("Delivery", info.delivery),
    ]

    row = 2
    for label, value in fields:
        ws.cell(row=row, column=1, value=label).font = label_font
        ws.cell(row=row, column=2, value=value)
        row += 1

    # ==========
    # Group Orders
    # ==========
    row += 1
    ws.cell(row=row, column=1, value="Group Orders").font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    ws.merge_cells(f"A{row}:D{row}")
    row += 1

    if result.group_orders:
        ws.cell(row=row, column=1, value="Group Order Number").font = label_font
        ws.cell(row=row, column=2, value=result.group_orders[0].group_order_number)
        row += 1

        ws.cell(row=row, column=1, value="Pick Time").font = label_font
        ws.cell(row=row, column=2, value=result.group_orders[0].pick_time)
        row += 1

    # =================
    # Individual Orders
    # =================
    row += 1
    ws.cell(row=row, column=1, value="Individual Orders").font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    ws.merge_cells(f"A{row}:H{row}")
    row += 1

    headers = [
        "Group Order #",
        "Guest Name",
        "Item Name",
        "Modification 1",
        "Modification 2",
        "Modification 3",
        "Modification 4",
        "Comments",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    row += 1

    for order in result.individual_orders:
        ws.cell(row=row, column=1, value=order.group_order_number)
        ws.cell(row=row, column=2, value=order.guest_name)
        ws.cell(row=row, column=3, value=order.item_name)

        # Modifications into separate columns (up to 4)
        for i, mod in enumerate(order.modifications[:4], start=4):
            ws.cell(row=row, column=i, value=mod)

        ws.cell(row=row, column=8, value=order.comments)
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 30
    for col in ["D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 25
    ws.column_dimensions["H"].width = 30

    wb.save(filepath)
    return str(filepath)
